import csv
import json
from typing import List, Dict
from datetime import datetime
import io


class DataExporter:
    """Export leads to various formats"""
    
    @staticmethod
    def to_csv(leads: List[Dict], include_metadata: bool = False) -> str:
        """Export leads to CSV format"""
        if not leads:
            return ""
        
        output = io.StringIO()
        
        base_columns = ['id', 'business_name', 'city', 'state', 'phone', 
                       'email', 'website', 'address', 'category', 'score', 'scraped_at']
        
        if include_metadata:
            base_columns.append('metadata')
        
        writer = csv.DictWriter(output, fieldnames=base_columns, extrasaction='ignore')
        writer.writeheader()
        
        for lead in leads:
            row = {k: lead.get(k, '') for k in base_columns}
            
            if include_metadata and 'metadata' in lead:
                row['metadata'] = json.dumps(lead['metadata'])
            
            writer.writerow(row)
        
        return output.getvalue()
    
    @staticmethod
    def to_json(leads: List[Dict], pretty: bool = True) -> str:
        """Export leads to JSON format"""
        for lead in leads:
            if 'scraped_at' in lead and isinstance(lead['scraped_at'], datetime):
                lead['scraped_at'] = lead['scraped_at'].isoformat()
        
        if pretty:
            return json.dumps(leads, indent=2, ensure_ascii=False)
        return json.dumps(leads, ensure_ascii=False)
    
    @staticmethod
    def to_excel_buffer(leads: List[Dict]):
        """Export to Excel format"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Leads"
            
            headers = ['ID', 'Business Name', 'City', 'State', 'Phone', 
                      'Email', 'Website', 'Address', 'Category', 'Score', 'Scraped At']
            ws.append(headers)
            
            for lead in leads:
                ws.append([
                    lead.get('id', ''),
                    lead.get('business_name', ''),
                    lead.get('city', ''),
                    lead.get('state', ''),
                    lead.get('phone', ''),
                    lead.get('email', ''),
                    lead.get('website', ''),
                    lead.get('address', ''),
                    lead.get('category', ''),
                    lead.get('score', ''),
                    str(lead.get('scraped_at', ''))
                ])
            
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
            
        except ImportError:
            raise ImportError("openpyxl required for Excel export")
